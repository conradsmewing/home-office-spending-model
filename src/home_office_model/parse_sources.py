"""Parsers that convert the raw gov.uk downloads into the tidy CSVs the model loads.

Run this module as a script to refresh ``data/raw/gdp_deflator.csv`` and
``data/raw/historic.csv`` from the files in ``data/raw/``.

    python -m home_office_model.parse_sources

Expected files in data/raw/:
    gdp_deflator.xlsx                HMT GDP deflators March 2026 QNA
    pesa_2025_ch1.xlsx               PESA 2025 Chapter 1 tables
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from .config import AREAS, DATA_RAW

HO_ROW_LABEL = "Home Office"


# ---------------------------------------------------------------------------
# GDP deflator
# ---------------------------------------------------------------------------

def parse_gdp_deflator(xlsx_path: Path | None = None, out_path: Path | None = None) -> pd.DataFrame:
    """Parse the HMT GDP deflator XLSX into a tidy CSV.

    The source sheet has the deflator rebased so that 2024-25 = 100 and covers
    1955-56 through 2030-31 (ONS outturn + OBR forecast).
    """
    xlsx_path = xlsx_path or DATA_RAW / "gdp_deflator.xlsx"
    out_path = out_path or DATA_RAW / "gdp_deflator.csv"

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []
    last_index = None
    for r in range(8, ws.max_row + 1):
        year = ws.cell(row=r, column=2).value
        idx = ws.cell(row=r, column=3).value
        pct = ws.cell(row=r, column=4).value
        if not year:
            continue
        year = str(year).strip()
        # Forecast years have a trailing "(1), (2)" annotation
        year_clean = year.split(" ")[0]
        if "-" not in year_clean or len(year_clean) != 7:
            continue
        if isinstance(idx, (int, float)):
            val = float(idx)
            source = "HMT March 2026 QNA (outturn)"
        elif isinstance(pct, (int, float)) and last_index is not None:
            # Forecast year — chain the % change onto last index
            val = last_index * (1 + float(pct) / 100.0)
            source = "HMT March 2026 QNA (OBR forecast)"
        else:
            continue
        last_index = val
        rows.append({"year": year_clean, "deflator_index": val, "source": source})

    df = pd.DataFrame(rows)
    df.to_csv(out_path, index=False)
    print(f"Wrote {len(df)} rows to {out_path} (last year: {df['year'].iloc[-1]})")
    return df


# ---------------------------------------------------------------------------
# PESA Chapter 1 — Home Office RDEL and CDEL
# ---------------------------------------------------------------------------

def _extract_ho_row(ws, year_row: int = 4, start_col: int = 2) -> dict[str, float]:
    """Given a PESA sheet, find the Home Office row and return {year: value}."""
    years = []
    for c in range(start_col, ws.max_column + 1):
        v = ws.cell(row=year_row, column=c).value
        if v and "-" in str(v):
            years.append((c, str(v).strip()))
    ho_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == HO_ROW_LABEL:
            ho_row = r
            break
    if ho_row is None:
        raise ValueError(f"Home Office row not found in sheet {ws.title}")
    out = {}
    for col, yr in years:
        val = ws.cell(row=ho_row, column=col).value
        if isinstance(val, (int, float)):
            out[yr] = float(val)
    return out


def parse_pesa_ho_totals(xlsx_path: Path | None = None) -> pd.DataFrame:
    """Extract HO Resource DEL (excl depreciation) and Capital DEL from PESA 2025 Ch1.

    Returns a tidy DataFrame: year, budget_type, value_gbp_m.
    Covers 2020-21 to 2025-26 (historic/plans) plus forward years from Tables 1.16 & 1.17.
    """
    xlsx_path = xlsx_path or DATA_RAW / "pesa_2025_ch1.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    rdel_hist = _extract_ho_row(wb["Table_1_5"])    # RDEL excl dep, 2020-21 to 2025-26
    rdel_fwd = _extract_ho_row(wb["Table_1_16"])    # RDEL excl dep, 2026-27 to 2028-29
    cdel_hist = _extract_ho_row(wb["Table_1_8"])    # CDEL, 2020-21 to 2025-26
    cdel_fwd = _extract_ho_row(wb["Table_1_17"])    # CDEL, 2026-27 to 2029-30

    rdel = {**rdel_hist, **rdel_fwd}
    cdel = {**cdel_hist, **cdel_fwd}

    records = []
    for yr, v in rdel.items():
        records.append({"year": yr, "budget_type": "RDEL", "value_gbp_m": v})
    for yr, v in cdel.items():
        records.append({"year": yr, "budget_type": "CDEL", "value_gbp_m": v})
    return pd.DataFrame(records).sort_values(["budget_type", "year"])


# ---------------------------------------------------------------------------
# Seed programme shares — applied to PESA totals to produce area breakdown
# ---------------------------------------------------------------------------

DEFAULT_RDEL_SHARES = {
    # Approximate shares based on HO ARA 2024-25 and published analysis. Replace
    # by parsing the HO ARA Statement of Parliamentary Supply for authoritative figures.
    "asylum_support": 0.190,         # ~£4.0bn of £21bn total DEL in 24-25
    "borders_migration": 0.098,
    "border_security_command": 0.005,
    "police": 0.608,
    "homeland_security": 0.041,
    "crime_fire_drugs": 0.035,
    "corporate_admin": 0.023,
}

DEFAULT_CDEL_SHARES = {
    "asylum_support": 0.04,
    "borders_migration": 0.22,
    "border_security_command": 0.02,
    "police": 0.32,
    "homeland_security": 0.18,
    "crime_fire_drugs": 0.04,
    "corporate_admin": 0.18,
}


def _check_shares_sum_to_one() -> None:
    for name, d in [("RDEL", DEFAULT_RDEL_SHARES), ("CDEL", DEFAULT_CDEL_SHARES)]:
        s = sum(d.values())
        if abs(s - 1.0) > 1e-6:
            raise ValueError(f"{name} shares sum to {s}, not 1.0")


_check_shares_sum_to_one()


def build_historic_from_pesa(
    pesa_totals: pd.DataFrame,
    rdel_shares: dict = None,
    cdel_shares: dict = None,
) -> pd.DataFrame:
    """Distribute PESA HO totals across programme areas using the share vectors above.

    The resulting DataFrame matches the schema of historic_seed.csv so it plugs straight
    into the existing loader. Years covered: 2020-21 to 2025-26 (historic + plans).
    """
    rdel_shares = rdel_shares or DEFAULT_RDEL_SHARES
    cdel_shares = cdel_shares or DEFAULT_CDEL_SHARES

    rows = []
    for _, row in pesa_totals.iterrows():
        yr = row["year"]
        bt = row["budget_type"]
        total = row["value_gbp_m"]
        shares = rdel_shares if bt == "RDEL" else cdel_shares
        for area in AREAS:
            rows.append({
                "year": yr,
                "area": area,
                "budget_type": bt,
                "value_gbp_m": total * shares[area],
                "source": f"PESA 2025 × default {bt} shares",
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# CLI entrypoint
# ---------------------------------------------------------------------------

def refresh_all() -> None:
    from .config import HISTORIC_END

    parse_gdp_deflator()
    pesa_totals = parse_pesa_ho_totals()
    pesa_totals.to_csv(DATA_RAW / "pesa_ho_totals.csv", index=False)

    historic = build_historic_from_pesa(pesa_totals)
    historic_outturn = historic[historic["year"] <= HISTORIC_END].copy()
    historic_outturn.to_csv(DATA_RAW / "historic.csv", index=False)
    print(
        f"Wrote historic.csv with {len(historic_outturn)} rows covering "
        f"{historic_outturn['year'].nunique()} outturn years"
    )

    # SR25 envelope (published plans for 2025-26 onwards)
    envelope = pesa_totals[pesa_totals["year"] > HISTORIC_END].copy()
    envelope.to_csv(DATA_RAW / "sr25_envelope.csv", index=False)
    print(f"Wrote sr25_envelope.csv with {len(envelope)} rows of PESA plans")


if __name__ == "__main__":
    refresh_all()
